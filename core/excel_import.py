"""Excel 匯入共用工具：封裝欄位取值、型別轉換、以及匯入流程樣板。"""

from datetime import datetime
from io import BytesIO

import pandas as pd
from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import redirect, render
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


class ImportRowError(Exception):
    """單列資料驗證失敗；外層會收集訊息並繼續處理下一列。"""


def get_value(row, candidates):
    """依序嘗試多個候選欄位名，回傳第一個有值者；皆無則回傳 None。"""
    if isinstance(candidates, str):
        candidates = [candidates]
    for col in candidates:
        if col in row.index:
            value = row[col]
            if pd.notna(value):
                return value
    return None


def clean_str(value, default=""):
    if value is None or pd.isna(value):
        return default
    return str(value).strip()


def require_str(value, field_name, row_num):
    text = clean_str(value)
    if not text:
        raise ImportRowError(f"第 {row_num} 列：{field_name}不可為空")
    return text


def parse_bool(value):
    if value is None or pd.isna(value):
        return False
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "t", "yes", "y", "是"}


def parse_date(value, default=None):
    if value is None or pd.isna(value):
        return default() if callable(default) else default
    return pd.to_datetime(value).date()


def today():
    return datetime.now().date()


def upsert(model, lookup, values, create_defaults=None):
    """僅用「非空值」更新既有資料；若不存在則以 values + create_defaults 建立。

    values 中為 None 或 "" 的欄位會被忽略（不會把既有資料洗成空）。
    create_defaults 用來補齊「新增時必填但 values 沒給」的欄位預設值。
    """
    creation = dict(create_defaults or {})
    for key, value in values.items():
        if value not in (None, ""):
            creation[key] = value

    obj, created = model.objects.get_or_create(**lookup, defaults=creation)
    if created:
        return obj

    changed = False
    for key, value in values.items():
        if value in (None, ""):
            continue
        if getattr(obj, key) != value:
            setattr(obj, key, value)
            changed = True
    if changed:
        obj.save()
    return obj


COMBINED_TEMPLATE_REQUIRED = ['員工姓名', '系統名稱', '登入帳號']
COMBINED_TEMPLATE_OPTIONAL = [
    '密碼',
    '所屬單位',
    '職稱',
    '到職日',
    '系統網址',
    '備註說明',
    '是否停用',
]
COMBINED_TEMPLATE_COLUMNS = COMBINED_TEMPLATE_REQUIRED + COMBINED_TEMPLATE_OPTIONAL


def build_combined_template_xlsx():
    """動態產生總表匯入用的 Excel 模板（只含標題列）。

    必填欄以深底白字標示；選填欄留原色，幫使用者一眼分辨。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "總表範例"
    ws.append(COMBINED_TEMPLATE_COLUMNS)

    required_fill = PatternFill(start_color="B4502E", end_color="B4502E", fill_type="solid")
    bold_white = Font(bold=True, color="FFFFFF")
    bold_dark = Font(bold=True)

    for idx, name in enumerate(COMBINED_TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx)
        if name in COMBINED_TEMPLATE_REQUIRED:
            cell.fill = required_fill
            cell.font = bold_white
        else:
            cell.font = bold_dark
        ws.column_dimensions[get_column_letter(idx)].width = max(len(name) * 2 + 4, 12)

    ws.freeze_panes = "A2"
    return wb


def combined_template_response(filename="系統帳密管理_總表範例.xlsx"):
    """把動態產生的模板包成 HttpResponse 回傳給瀏覽器下載。"""
    wb = build_combined_template_xlsx()
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def handle_excel_import(request, *, template_name, entity_label, row_handler):
    """通用 Excel 匯入流程。

    - GET：渲染上傳表單。
    - POST：讀取檔案、逐列交給 row_handler 處理、彙整訊息後重導。

    row_handler(row, row_num) 需自行負責建檔/更新；遇列錯誤請 raise ImportRowError。
    """
    if request.method != "POST":
        return render(request, template_name)

    excel_file = request.FILES.get("excel_file")
    if not excel_file:
        messages.error(request, "請選擇一個 Excel 檔案！")
        return redirect("..")

    try:
        df = pd.read_excel(excel_file)
    except Exception as e:
        messages.error(request, f"匯入失敗，請檢查 Excel 格式。錯誤訊息：{e}")
        return redirect("..")

    count = 0
    errors = []
    for index, row in df.iterrows():
        row_num = index + 2  # Excel 第 1 列為標題
        try:
            row_handler(row, row_num)
        except ImportRowError as e:
            errors.append(str(e))
        except Exception as e:
            errors.append(f"第 {row_num} 列：未預期錯誤：{e}")
        else:
            count += 1

    if errors:
        messages.error(request, "\n".join(errors))
    if count:
        messages.success(request, f"成功匯入了 {count} 筆{entity_label}資料！🎉")
    elif not errors:
        messages.warning(request, "Excel 中沒有可匯入的資料。")

    return redirect("..")
